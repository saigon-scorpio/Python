import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook


#Needed those packet installation in command prompt in windows or terminal in mac
#pip or pip3 install requests and pip3 install beautifulsoup4
#pip install openpyxl


def pronounciation (w):
    url = 'https://www.merriam-webster.com/dictionary/'+w
    r = requests.get(url)
    soup = BeautifulSoup(r.content,"lxml")
    section = soup.find_all('span', class_='pr')
    pr =[]
    for i in section:
        pr.append(i.find(text=True, recursive=False))
    if len(pr) == 0:
        pr.append('Word doesn\'t exist in dictionary')
    return pr

def moving(a,b):
    
    for i in range(len(b)):
        a.append(b.pop(0))

def main():
  
  bookOutput = Workbook()
  sheetOutput = bookOutput.active
  sheetOutput.cell(row=1, column=1).value = 'Word'
  sheetOutput.cell(row=1, column=2).value = 'Pronounciation'

# name of the excel file that HAVE TO BE in the same folder with python
  bookInput = load_workbook("test.xlsx")
  shellInput = bookInput.active
  columns = shellInput['A']
  
  for i in range(len(columns)):
      row = [columns[i].value]
      moving(row,pronounciation(columns[i].value))
      for j in range(len(row)):
          sheetOutput.cell(row=(i+2), column=(j+1)).value = str(row[j])

# This is the name of the output file
  bookOutput.save('output.xlsx')   
main()
    
